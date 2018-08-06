from openpyxl import *
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color,NamedStyle
from FormulesCouts import *
from FormulesRevenus import *
import pandas as pd
import numpy as np
import os



def export_dfs(df_list, sheet, file_name, spaces,colstart=0):
    writer = pd.ExcelWriter(file_name,engine='xlsxwriter')
    row = 4
    for dataframe in df_list:
        dataframe.to_excel(writer,sheet_name=sheet,startrow=row , startcol=colstart)   
        row = row + len(dataframe.index) + spaces + 1
    writer.save()

def excel_dfs(Dict, file_name, spaces):
    fontsDict = {'Activite':Font(name='Calibri',size=16,bold = True),
                 'Cout':Font(name='Calibri',size=11,bold=True,italic=True),
                 'Revenu':Font(name='Calibri',size=11,bold=True,italic=True),
                 'Titre':Font(name='Calibri',size=11,italic=True)}
    indexColAct = 0
    WordRowCol = {}
    DictRecolte = {}
    writer = pd.ExcelWriter(file_name,engine='xlsxwriter')
    sheets = ['Int couts','Mar couts','Int revenus','Mar revenus']
    #Début ecriture données
    for sheet in sheets:
        indexAct = 0
        WordRowCol[sheet] = []
        DictRecolte[sheet] = {}
        pasActivite = 0
        Listespaces = [0]
        for activite in Dict[sheet].keys():
            Listespaces.append(max([a[1] for a in Dict[sheet][activite].values()])+3)
        for activite in Dict[sheet].keys():
            indexAct+=max(Listespaces)
            row = 5
            WordRowCol[sheet].append((activite,row-4,indexAct+2,fontsDict['Activite']))
            DictRecolte[sheet][activite] = {}
            for cout in Dict[sheet][activite].keys():
                if Dict[sheet][activite][cout][0] != []:
                    WordRowCol[sheet].append((cout,row-2,indexAct,fontsDict['Cout']))
                    DictRecolte[sheet][activite][cout] = {}
                    for dataframe in Dict[sheet][activite][cout][0]:
                        WordRowCol[sheet].append((dataframe[1],row+1,indexAct+1,fontsDict['Titre']))
                        DictRecolte[sheet][activite][cout][dataframe[1]] = [row+1,indexAct+1]
                        dataframe[0].to_excel(writer,sheet_name=sheet,startrow=row , startcol=indexAct)   
                        row = row + len(dataframe[0].index) + spaces + 1
    writer.save()
    #Fin ecritre données
    wb = load_workbook('Parametres.xlsx')
    # for sheet in WordRowCol.keys():
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for i in WordRowCol[sheet]:
            ws.cell(i[1],i[2]).value = i[0]
            ws.cell(i[1],i[2]).font = i[3]
        for col in ws.columns:
            max_length = 0
            column = col[0].column # Get the column name
            for cell in col:
                try: # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width
    wb.save('Parametres.xlsx')
    return WordRowCol,DictRecolte
            

    

class Projet:
    def __init__(self,Nom="",Localisation="",ListeActivites=[],Horizon=0,CPC=[]):
        self.Nom = Nom
        self.ListeActivitesSecondaires = []
        self.Localisation = Localisation
        self.ListeActivites = ListeActivites
        self.Horizon = Horizon
        self.CPC = CPC
        self.pasVisualisation = ""
        self.DictParams = {"Int couts":{},
                           "Mar couts":{},
                           "Int revenus":{},
                           "Mar revenus":{}}
        self.WordRowCol = {}
        self.DictRecolte = {"Int couts":{},
                           "Mar couts":{}}
        
        self.ProduitsFinanciers = []
        self.ChargesFinancieres = []
        self.Interets = []
        self.Amortissements = []
        self.RAI = []
        self.Impots = []
        self.ResultatNet = []
        self.FondsPropres = []
        self.CCA = []
        self.Dette = []
        self.RemboursementDette = []
        self.FluxTresorerie = []
        self.FluxTresorerieAcc = []
        self.DetteObj = []
        self.ListeInvest = []
        
    def InitialiserDette(self):
        self.DetteObj = Dette()
        
    def distibuerPas(self):
        for activite in self.ListeActivites:
            for cout in activite.getlistCout():
                cout.pasVisualisation = self.pasVisualisation
            for revenu in activite.getlistRev():
                revenu.pasVisualisation = self.pasVisualisation
    
    def PrepareExcelInput(self):
        for activite in self.ListeActivites:
            for key in self.DictParams.keys():
                self.DictParams[key][activite.getNom()] = {}
            for cout in activite.getlistCout():
                MaxCols = max([i.getTableauAffichage().shape[1] for i in cout.getListTableaux()+cout.getListTableauxMarche()])
                # Attention au changement du titre
                self.DictParams['Int couts'][activite.getNom()][cout.getNom()] = ([(T.getTableauAffichage(),T.getTitre()) for T in cout.getListTableaux()],MaxCols) 
                self.DictParams['Mar couts'][activite.getNom()][cout.getNom()] = ([(T.getTableauAffichage(),T.getTitre()) for T in cout.getListTableauxMarche()],MaxCols)
        
            
            for revenu in activite.getlistRev():
                MaxColsRev =  max([i.getTableauAffichage().shape[1] for i in revenu.getListTableaux()+revenu.getListTableauxMarche()])
                self.DictParams['Int revenus'][activite.getNom()][revenu.getNom()] = ([(T.getTableauAffichage(),T.getTitre()) for T in revenu.getListTableaux()],MaxColsRev)
                self.DictParams['Mar revenus'][activite.getNom()][revenu.getNom()] = ([(T.getTableauAffichage(),T.getTitre()) for T in revenu.getListTableauxMarche()],MaxColsRev)
        self.WordRowCol,self.DictRecolte = excel_dfs(self.DictParams,"Parametres.xlsx",4)
    
    def GetExcelInput(self):
        wb = load_workbook('Parametres.xlsx',data_only=True)
        # wsIC = wb['Int couts']
        # wsMC = wb['Mar couts']
        # wsIR = wb['Int revenus']
        # wsMR = wb['Mar revenus']
        for activite in self.ListeActivites:
            for cout in activite.getlistCout():
                if len(cout.getListTableaux()) != 0:
                    for Tableau in cout.getListTableaux():
                        CornerRow = self.DictRecolte['Int couts'][activite.getNom()][cout.getNom()][Tableau.getTitre()][0]+1
                        CornerColumn = self.DictRecolte['Int couts'][activite.getNom()][cout.getNom()][Tableau.getTitre()][1]+1
                        Tableau.setTableauDonnees(np.array([[wb['Int couts'].cell(i,j).value for j in range(CornerColumn,CornerColumn + Tableau.Taille[1])] for i in range(CornerRow,CornerRow + Tableau.Taille[0])]))
                if len(cout.getListTableauxMarche()) != 0:
                    for Tableau in cout.getListTableauxMarche():
                        CornerRow = self.DictRecolte['Mar couts'][activite.getNom()][cout.getNom()][Tableau.getTitre()][0]+1
                        CornerColumn = self.DictRecolte['Mar couts'][activite.getNom()][cout.getNom()][Tableau.getTitre()][1]+1
                        Tableau.setTableauDonnees(np.array([[wb['Mar couts'].cell(i,j).value for j in range(CornerColumn,CornerColumn + Tableau.Taille[1])] for i in range(CornerRow,CornerRow + Tableau.Taille[0])]))
                        ######################
            for revenu in activite.getlistRev():
                if len(revenu.getListTableaux()) != 0:
                    for Tableau in revenu.getListTableaux():
                        CornerRow = self.DictRecolte['Int revenus'][activite.getNom()][revenu.getNom()][Tableau.getTitre()][0]+1
                        CornerColumn = self.DictRecolte['Int revenus'][activite.getNom()][revenu.getNom()][Tableau.getTitre()][1]+1
                        Tableau.setTableauDonnees(np.array([[wb['Int revenus'].cell(i,j).value for j in range(CornerColumn,CornerColumn + Tableau.Taille[1])] for i in range(CornerRow,CornerRow + Tableau.Taille[0])]))
                if len(revenu.getListTableauxMarche()) != 0:   
                    for Tableau in revenu.getListTableauxMarche():
                        CornerRow = self.DictRecolte['Mar revenus'][activite.getNom()][revenu.getNom()][Tableau.getTitre()][0]+1
                        CornerColumn = self.DictRecolte['Mar revenus'][activite.getNom()][revenu.getNom()][Tableau.getTitre()][1]+1
                        Tableau.setTableauDonnees(np.array([[wb['Mar revenus'].cell(i,j).value for j in range(CornerColumn,CornerColumn + Tableau.Taille[1])] for i in range(CornerRow,CornerRow + Tableau.Taille[0])]))
                    

    
    #Méthode pour enregistrer tout les tableaux
        
    def IdentifierActivitesPossibles(self):
        self.ListeActivites,self.ListeActivitesSecondaires = [],[]
        print("Identification des Activités")
        wb = load_workbook('References.xlsx')
        ws = wb['Ref activite']
        for col in ws.iter_cols(min_row=2,min_col=3, max_col=ws.max_column, max_row=2):
            for cell in col:
                if cell.value == self.getNom():
                    col_to_search = cell.col_idx
        for row in ws.iter_rows(min_row=3,max_row=ws.max_row,min_col=col_to_search,max_col=col_to_search):
            for cell in row:
                if cell.value == 1:
                    A = Activite()
                    A.setNom(ws['C'+str(cell.row)].value)
                    self.ListeActivites.append(A)
                elif cell.value == 2:
                    A = Activite()
                    A.setNom(ws['C'+str(cell.row)].value)
                    self.ListeActivitesSecondaires.append(A)
                    
        
    def AjoutActivites(self,NewListeActiv):
        self.ListeActivites += NewListeActiv
    
    def CommencerSaisie(self):
        pass
    def GenerateTFT(self):
        wb = Workbook()
        wb.create_sheet("TFT")
        ws=wb['TFT']
        style1 = NamedStyle(name="style1")
        style1.font = Font(name="Calibri",bold="True",size=13,color="FFFFFF")
        style1.alignment = Alignment(horizontal = "center", vertical = "center")
        style1.fill = PatternFill("solid",fgColor="C00000")
        # Style 2 - PRODUITS D'EXPLOITATION
        style2 = NamedStyle(name="style2")
        style2.font = Font(name="Calibri",bold="True",size=13,color="000000")
        style2.alignment = Alignment(horizontal = "left", vertical = "center")
        style2.fill = PatternFill("solid",fgColor="AEAAAA")
        # Style 2 - PRODUITS D'EXPLOITATION
        style2_numbers = NamedStyle(name="style2_numbers")
        style2_numbers.font = Font(name="Calibri",bold="True",size=11,color="000000")
        style2_numbers.alignment = Alignment(horizontal = "center", vertical = "center",indent = 0)
        style2_numbers.fill = PatternFill("solid",fgColor="AEAAAA")
        # Style 3 - Prestation de cours 
        style3 = NamedStyle(name="style3")
        style3.font = Font(name="Calibri",bold="True",size=11,color="000000")
        style3.alignment = Alignment(horizontal = "left", vertical = "center")
        style3.fill = PatternFill("solid",fgColor="D9D9D9")
        # Style 3_numbers - Prestation de cours 
        style3_numbers = NamedStyle(name="style3_numbers")
        style3_numbers.font = Font(name="Calibri",bold="True",size=11,color="000000")
        style3_numbers.alignment = Alignment(horizontal = "center", vertical = "center",indent=0)
        style3_numbers.fill = PatternFill("solid",fgColor="D9D9D9")
        # Style 4 - Inscriptions enregistrés
        style4 = NamedStyle(name="style4")
        style4.font = Font(name="Calibri",bold="True",size=10,color="000000")
        style4.alignment = Alignment(horizontal = "left", vertical = "center",indent = 5)
        style4.fill = PatternFill("solid",fgColor="D9D9D9")
        # Style 4_Numbers
        style4_numbers = NamedStyle(name="style4_numbers")
        style4_numbers.font = Font(name="Calibri",bold="True",size=11,color="000000")
        style4_numbers.alignment = Alignment(horizontal = "center", vertical = "center",indent = 0)
        style4_numbers.fill = PatternFill("solid",fgColor="D9D9D9")
        # Style 5 - TOTAL
        style5 = NamedStyle(name="style5")
        style5.font = Font(name="Calibri",bold="True",size=13,color="000000")
        style5.alignment = Alignment(horizontal = "left", vertical = "center",indent=0)
        style5.fill = PatternFill("solid",fgColor="AEAAAA")
        Thin = Side(border_style="thin", color="000000")
        style5.border = Border(bottom = Thin)
        #Style 5 Total_numbers
        style5_numbers = NamedStyle(name="style5_numbers")
        style5_numbers.font = Font(name="Calibri",bold="True",size=11,color="000000")
        style5_numbers.alignment = Alignment(horizontal = "center", vertical = "center",indent=0)
        style5_numbers.fill = PatternFill("solid",fgColor="AEAAAA")
        Thin = Side(border_style="thin", color="000000")
        style5_numbers.border = Border(bottom = Thin)
        # Style 6 - RESULTAT D'EXPLOITATION
        style6 = NamedStyle(name="style6")
        style6.font = Font(name="Calibri",bold="True",size=13,color="FFFFFF")
        style6.alignment = Alignment(horizontal = "left", vertical = "center")
        style6.fill = PatternFill("solid",fgColor="C00000")
        #Style 6 Numbers
        style6_numbers = NamedStyle(name="style6_numbers")
        style6_numbers.font = Font(name="Calibri",bold="True",size=11,color="FFFFFF")
        style6_numbers.alignment = Alignment(horizontal = "center", vertical = "center",indent = 0)
        style6_numbers.fill = PatternFill("solid",fgColor="C00000")
        ws['A1'].value= 'TFT Projet '+self.Nom
        ws['A1'].style = 'Headline 1'
        for i in range(self.Horizon):
            ws.cell(3,i+2).value = self.pasVisualisation+str(i+1)
            ws.cell(3,i+2).style = style1
       
        PTFT = 4
        ws.cell(PTFT,1).value = "Flux du Projet"
        ws.cell(PTFT,1).style = style2
        for i in range(self.Horizon):
            ws.cell(PTFT,i+2).value = ""
            ws.cell(PTFT,i+2).style = style2
        PTFT += 1
        #### Resultat Avant Impot
        ws.cell(PTFT,1).value = "Resultat Avant Impôts"
        ws.cell(PTFT,1).style = style4
        for i in range(self.Horizon):
            ws.cell(PTFT,i+2).value = self.RAI[i]
            ws.cell(PTFT,i+2).style = style4_numbers
        PTFT += 1
        # # # # # IS
        ws.cell(PTFT,1).value = "IS"
        ws.cell(PTFT,1).style = style4
        for i in range(self.Horizon):
            ws.cell(PTFT,i+2).value = self.Impots[i]
            ws.cell(PTFT,i+2).style = style4_numbers
        PTFT += 1
        # # # # # Ammortissement
        self.Amortissement = [0]*self.Horizon
        ws.cell(PTFT,1).value = "Amortissement"
        ws.cell(PTFT,1).style = style4
        for i in range(self.Horizon):
            ws.cell(PTFT,i+2).value = self.Amortissement[i]
            ws.cell(PTFT,i+2).style = style4_numbers
        PTFT += 1
        # # # # # Variation BFR
        self.VBFR = [0]*self.Horizon
        ws.cell(PTFT,1).value = "Variation BFR"
        ws.cell(PTFT,1).style = style4
        for i in range(self.Horizon):
            ws.cell(PTFT,i+2).value = self.VBFR[i]
            ws.cell(PTFT,i+2).style = style4_numbers
        PTFT += 1
        ##### Flux Invesstissement
        ws.cell(PTFT,1).value = "Flux d'Invesstissement"
        ws.cell(PTFT,1).style = style2
        for i in range(self.Horizon):
            ws.cell(PTFT,i+2).value = ""
            ws.cell(PTFT,i+2).style = style2
        PTFT += 1
        for i in range(len(self.ListeInvest)):
            ws.cell(PTFT,1).value = self.ListeInvest[i][0]
            ws.cell(PTFT,1).style = style4
            for j in range(1,self.Horizon+1):
                if j < len(self.ListeInvest[i]):
                    ws.cell(PTFT,j+1).value = self.ListeInvest[i][j]
                    ws.cell(PTFT,j+1).style = style4_numbers 
                else:
                    ws.cell(PTFT,j+1).value = 0
                    self.ListeInvest[i].append(0)
                    ws.cell(PTFT,j+1).style = style4_numbers 
            PTFT += 1

        ##### Flux de financement
        ws.cell(PTFT,1).value = "Flux de financement"
        ws.cell(PTFT,1).style = style2
        for i in range(self.Horizon):
            ws.cell(PTFT,i+2).value = ""
            ws.cell(PTFT,i+2).style = style2
        PTFT += 1
        # # # # # Fonds Propres Injectés
        l = [0]*self.Horizon
        ws.cell(PTFT,1).value = "Fond Propres"
        ws.cell(PTFT,1).style = style4
        for i in range(len(self.FondsPropres)):
            l[i] = self.FondsPropres[i]
        self.FondsPropres = l.copy()
        for i in range(self.Horizon):
            ws.cell(PTFT,i+2).value = self.FondsPropres[i]
            ws.cell(PTFT,i+2).style = style4_numbers
        
        PTFT += 1
        # # # # # Levée de Dette
        l = [0]*self.Horizon
        ws.cell(PTFT,1).value = "Levée de Dette"
        ws.cell(PTFT,1).style = style4
        
        for i in range(len(self.Dette)):
            l[i] = self.Dette[i]
        self.Dette = l.copy()
        for i in range(self.Horizon):
            ws.cell(PTFT,i+2).value = self.Dette[i]
            ws.cell(PTFT,i+2).style = style4_numbers
        
        PTFT += 1
        # # # # Apport en CCA
        ws.cell(PTFT,1).value = "Apport en CCA"
        ws.cell(PTFT,1).style = style4
        for i in range(len(self.CCA)):
            l[i] = self.CCA[i]
        self.CCA = l.copy()
        for i in range(self.Horizon):
            ws.cell(PTFT,i+2).value = self.CCA[i]
            ws.cell(PTFT,i+2).style = style4_numbers
        
        PTFT += 1
        # # # # # Remboursement de la dette (Capital)
        self.RemboursementDette = [0]*self.Horizon
        ws.cell(PTFT,1).value = "Remboursement de la dette"
        ws.cell(PTFT,1).style = style4
        for i in range(self.Horizon):
            ws.cell(PTFT,i+2).value = self.RemboursementDette[i]
            ws.cell(PTFT,i+2).style = style4_numbers
        PTFT += 1
        # # # # # Remboursement du CCA
        self.RemboursementCCA = [0]*self.Horizon
        ws.cell(PTFT,1).value = "Remboursement CCA"
        ws.cell(PTFT,1).style = style4
        for i in range(self.Horizon):
            ws.cell(PTFT,i+2).value = self.RemboursementCCA[i]
            ws.cell(PTFT,i+2).style = style4_numbers
        PTFT += 1
        # # # # # Distribution de dividendes
        self.RemboursementDividendes = [0]*self.Horizon
        ws.cell(PTFT,1).value = "Distribution Dividendes"
        ws.cell(PTFT,1).style = style4
        for i in range(self.Horizon):
            ws.cell(PTFT,i+2).value = self.RemboursementDividendes[i]
            ws.cell(PTFT,i+2).style = style4_numbers
        PTFT += 1
        # # # # # # Interets financiers
        # ws.cell(PTFT,1).value = "Interets financiers"
        # ws.cell(PTFT,1).style = style3
        # for i in range(self.Horizon):
        #     ws.cell(PTFT,i+2).value = self.Interets[i]
        #     ws.cell(PTFT,i+2).style = style3_numbers
        # PTFT += 1
        # # # # # Flux de trésorerie
        ws.cell(PTFT,1).value = "Flux de trésorerie"
        ws.cell(PTFT,1).style = style2
        self.FluxTresorerie = [0]*self.Horizon
        Investissements = [sum(item) for item in zip(*[i[1:] for i in self.ListeInvest])]
        if len(Investissements) !=0:
            for i in range(self.Horizon):
                self.FluxTresorerie[i] = self.RAI[i]-self.Impots[i]+self.Amortissements[i]+self.VBFR[i]-Investissements[i]+self.FondsPropres[i]+self.Dette[i]+self.CCA[i]-self.RemboursementDette[i]-self.RemboursementCCA[i]-self.RemboursementDividendes[i]
                ws.cell(PTFT,i+2).value = self.FluxTresorerie[i]
                ws.cell(PTFT,i+2).style = style2_numbers
            PTFT += 1
        else:
            for i in range(self.Horizon):
                self.FluxTresorerie[i] = self.RAI[i]-self.Impots[i]+self.Amortissements[i]+self.VBFR[i]+self.FondsPropres[i]+self.Dette[i]+self.CCA[i]-self.RemboursementDette[i]-self.RemboursementCCA[i]-self.RemboursementDividendes[i]
                ws.cell(PTFT,i+2).value = self.FluxTresorerie[i]
                ws.cell(PTFT,i+2).style = style2_numbers
            PTFT += 1
        # # # # # Trésorerie accumulée
        ws.cell(PTFT,1).value = "Trésorerie Accumulée"
        ws.cell(PTFT,1).style = style2
        self.FluxTresorerieAcc = [0]*self.Horizon
        for i in range(self.Horizon):
            if i == 0:
                self.FluxTresorerieAcc[i] = self.FluxTresorerie[i]
                ws.cell(PTFT,i+2).value = self.FluxTresorerieAcc[i]
            else:
                self.FluxTresorerieAcc[i] = self.FluxTresorerieAcc[i-1] + self.FluxTresorerie[i]
                ws.cell(PTFT,i+2).value = self.FluxTresorerieAcc[i]
            ws.cell(PTFT,i+2).style = style2_numbers
        PTFT += 1
        wb.remove(wb['Sheet'])
        wb.save("tft.xlsx")
    def GenerateCPC(self):
        ######### Partie calcul des éléments du cpc ##########
        ######################################################
        # Style 1 --- A1,A2----An
        style1 = NamedStyle(name="style1")
        style1.font = Font(name="Calibri",bold="True",size=13,color="FFFFFF")
        style1.alignment = Alignment(horizontal = "center", vertical = "center")
        style1.fill = PatternFill("solid",fgColor="C00000")
        # Style 2 - PRODUITS D'EXPLOITATION
        style2 = NamedStyle(name="style2")
        style2.font = Font(name="Calibri",bold="True",size=13,color="000000")
        style2.alignment = Alignment(horizontal = "left", vertical = "center")
        style2.fill = PatternFill("solid",fgColor="AEAAAA")
        # Style 2 - PRODUITS D'EXPLOITATION
        style2_numbers = NamedStyle(name="style2_numbers")
        style2_numbers.font = Font(name="Calibri",bold="True",size=11,color="000000")
        style2_numbers.alignment = Alignment(horizontal = "center", vertical = "center",indent = 0)
        style2_numbers.fill = PatternFill("solid",fgColor="AEAAAA")
        # Style 3 - Prestation de cours 
        style3 = NamedStyle(name="style3")
        style3.font = Font(name="Calibri",bold="True",size=11,color="000000")
        style3.alignment = Alignment(horizontal = "left", vertical = "center")
        style3.fill = PatternFill("solid",fgColor="D9D9D9")
        # Style 4 - Inscriptions enregistrés
        style4 = NamedStyle(name="style4")
        style4.font = Font(name="Calibri",bold="True",size=10,color="000000")
        style4.alignment = Alignment(horizontal = "left", vertical = "center",indent = 5)
        style4.fill = PatternFill("solid",fgColor="D9D9D9")
        # Style 4_Numbers
        style4_numbers = NamedStyle(name="style4_numbers")
        style4_numbers.font = Font(name="Calibri",bold="True",size=11,color="000000")
        style4_numbers.alignment = Alignment(horizontal = "center", vertical = "center",indent = 0)
        style4_numbers.fill = PatternFill("solid",fgColor="D9D9D9")
        # Style 5 - TOTAL
        style5 = NamedStyle(name="style5")
        style5.font = Font(name="Calibri",bold="True",size=13,color="000000")
        style5.alignment = Alignment(horizontal = "left", vertical = "center",indent=0)
        style5.fill = PatternFill("solid",fgColor="AEAAAA")
        Thin = Side(border_style="thin", color="000000")
        style5.border = Border(bottom = Thin)
        #Style 5 Total_numbers
        style5_numbers = NamedStyle(name="style5_numbers")
        style5_numbers.font = Font(name="Calibri",bold="True",size=11,color="000000")
        style5_numbers.alignment = Alignment(horizontal = "center", vertical = "center",indent=0)
        style5_numbers.fill = PatternFill("solid",fgColor="AEAAAA")
        Thin = Side(border_style="thin", color="000000")
        style5_numbers.border = Border(bottom = Thin)
        # Style 6 - RESULTAT D'EXPLOITATION
        style6 = NamedStyle(name="style6")
        style6.font = Font(name="Calibri",bold="True",size=13,color="FFFFFF")
        style6.alignment = Alignment(horizontal = "left", vertical = "center")
        style6.fill = PatternFill("solid",fgColor="C00000")
        
        #Style 6 Numbers
        style6_numbers = NamedStyle(name="style6_numbers")
        style6_numbers.font = Font(name="Calibri",bold="True",size=11,color="FFFFFF")
        style6_numbers.alignment = Alignment(horizontal = "center", vertical = "center",indent = 0)
        style6_numbers.fill = PatternFill("solid",fgColor="C00000")
        
        #cette méthode permet de genere un cpc dans un premier temps en vision anuelle 
        #Section 1 : Ouverture et definition de la taille
        wb = Workbook()
        wb.create_sheet("CPC")
        ws=wb['CPC']
        TotalRevenusOp = [0]*self.Horizon
        TotalCoutsOp= [0]*self.Horizon
        ws['A1'] = 'CPC Projet '+self.Nom
        ws['A1'].style = 'Headline 1'
        ws['A4'] = 'PRODUITS D\'EXPLOITATION'
        ws['A4'].style = style2
        for i in range(self.Horizon):
            ws.cell(4,i+2).style = style2
        Horizon = self.Horizon
        for i in range(Horizon):
            ws.cell(3,i+2).value = self.pasVisualisation+str(i+1)
            ws.cell(3,i+2).style = style1 
        ####Section Produits d'exploitation
        PCPC = 5
        for activite in self.ListeActivites:
            ws['A'+str(PCPC)].value = activite.getNom()
            ws['A'+str(PCPC)].style = style3
            for i in range(Horizon):
                ws.cell(PCPC,i+2).style = style3
            PCPC += 1
            for revenu in activite.getlistRev():
                ws['A'+str(PCPC)].value = revenu.getNom()
                ws['A'+str(PCPC)].style = style4
                for i in range(Horizon):
                    ws.cell(PCPC,i+2).value = revenu.resultat[i]
                    ws.cell(PCPC,i+2).style = style4_numbers
                    TotalRevenusOp[i]+=revenu.resultat[i]
                PCPC +=1
        ws['A'+str(PCPC)].value = 'TOTAL PRODUITS D\'EXPLOITATION'
        ws['A'+str(PCPC)].style = style5
        for i in range(Horizon):
            ws.cell(PCPC,i+2).value = TotalRevenusOp[i]
            ws.cell(PCPC,i+2).style = style5_numbers
        PCPC +=1
        ### Section Charges Exploitation
        ws['A'+str(PCPC)].value = 'CHARGES D\'EXPLOITATION'
        ws['A'+str(PCPC)].style = style2
        for i in range(self.Horizon):
            ws.cell(PCPC,i+2).style = style2
        PCPC +=1
        for activite in self.ListeActivites:
            ws['A'+str(PCPC)].value = activite.getNom()
            ws['A'+str(PCPC)].style = style3
            for i in range(self.Horizon):
                ws.cell(PCPC,i+2).style = style3
            PCPC += 1
            for cout in activite.getlistCout():
                ws['A'+str(PCPC)].value = cout.getNom()
                ws['A'+str(PCPC)].style = style4
                for i in range(Horizon):
                    ws.cell(PCPC,i+2).value = cout.resultat[i]
                    ws.cell(PCPC,i+2).style = style4_numbers
                    TotalCoutsOp[i]+=cout.resultat[i]
                PCPC +=1
        ws['A'+str(PCPC)].value = 'TOTAL CHARGES D\'EXPLOITATION'
        ws['A'+str(PCPC)].style = style5
        for i in range(Horizon):
            ws.cell(PCPC,i+2).value = TotalCoutsOp[i]
            ws.cell(PCPC,i+2).style = style5_numbers
        PCPC +=1
        ws['A'+str(PCPC)].value = 'RESULTAT D\'EXPLOITATION'
        ws['A'+str(PCPC)].style = style6
        TotalResultatOp = [0]*self.Horizon
        for i in range(Horizon):
            TotalResultatOp[i] = TotalRevenusOp[i] - TotalCoutsOp[i]
            ws.cell(PCPC,i+2).value = TotalResultatOp[i]
            ws.cell(PCPC,i+2).style = style6_numbers
        PCPC +=1
        ###Section Produits Financiers
        ws['A'+str(PCPC)].value = 'PRODUITS FINANCIERS'
        ws['A'+str(PCPC)].style = style2
        self.ProduitsFinanciers = [0]*self.Horizon
        for i in range(self.Horizon):
            ws.cell(PCPC,i+2).value = self.ProduitsFinanciers[i]
            ws.cell(PCPC,i+2).style = style2_numbers
        PCPC +=1
        ###Section Charges Financières
        ws['A'+str(PCPC)].value = 'CHARGES FINANCIERES'
        ws['A'+str(PCPC)].style = style2
        self.Interets = [0]*self.Horizon
        for i in range(self.Horizon):
            ws.cell(PCPC,i+2).style = style2
        PCPC +=1
        ws['A'+str(PCPC)].value = 'Intêrets'
        ws['A'+str(PCPC)].style = style4
        for i in range(self.Horizon):
            ws.cell(PCPC,i+2).value = self.Interets[i]
            ws.cell(PCPC,i+2).style = style4_numbers
        PCPC +=1
        ###Section Resultat Financiers
        ws['A'+str(PCPC)].value = 'RESULTAT FINANCIER'
        ws['A'+str(PCPC)].style = style6
        ResultatFinancier = [0]*self.Horizon
        for i in range(self.Horizon):
            ResultatFinancier[i] = self.ProduitsFinanciers[i] - self.Interets[i]
            ws.cell(PCPC,i+2).style = style6_numbers
        PCPC += 1
        ###Section Dotations
        ws['A'+str(PCPC)].value = 'DOTATIONS'
        ws['A'+str(PCPC)].style = style2
        for i in range(self.Horizon):
            ws.cell(PCPC,i+2).style = style2
        PCPC += 1
        ws['A'+str(PCPC)].value = 'Amortissements'
        ws['A'+str(PCPC)].style = style4
        self.Amortissements = [0]*self.Horizon
        for i in range(self.Horizon):
            ws.cell(PCPC,i+2).value = self.Amortissements[i]
            ws.cell(PCPC,i+2).style = style4_numbers
        PCPC +=1
        
        ###Section Resultat avant Impots
        ws['A'+str(PCPC)].value = 'RESULTAT AVANT IMPOT'
        ws['A'+str(PCPC)].style = style6
        self.RAI =[0]*self.Horizon
        for i in range(self.Horizon):
            self.RAI[i] = TotalResultatOp[i] + ResultatFinancier[i] - self.Amortissements[i]
            ws.cell(PCPC,i+2).value = self.RAI[i]
            ws.cell(PCPC,i+2).style =  style6_numbers
        PCPC += 1
        
        ###Section Impôts
        ws['A'+str(PCPC)].value = 'IMPÔTS'
        ws['A'+str(PCPC)].style = style2
        self.Impots = IS(self)
        for i in range(self.Horizon):
            ws.cell(PCPC,i+2).value = self.Impots[i]
            ws.cell(PCPC,i+2).style = style2_numbers
        PCPC += 1 
           
        ###Section Resultats nets
        ws['A'+str(PCPC)].value = 'RESULTAT NET'
        ws['A'+str(PCPC)].style = style6
        self.ResultatNet = [0]*self.Horizon
        for i in range(self.Horizon):
            self.ResultatNet[i] = self.RAI[i] - self.Impots[i]
            ws.cell(PCPC,i+2).value = self.ResultatNet[i]
            ws.cell(PCPC,i+2).style = style6_numbers
        wb.remove(wb['Sheet'])
        wb.save("cpc.xlsx")

    #Setters 
    def setNom(self,Nom):
        self.Nom = Nom
    def setLocalisation(self,Localisation):
        self.Localisation = Localisation
    def setListeActivite(self,ListeActivites):
        self.ListeActivites = ListeActivites
    def setHorizon(self,Horizon):
        self.Horizon = Horizon
    def setDictParams(self,Dict):
        self.DictParams = Dict

    #Getters
    def getNom(self):
        return self.Nom
    def getLocalisation(self):
        return self.Localisation
    def getListeActivites(self):
        return self.ListeActivites
    def showListeActivites(self):
        for i in self.ListeActivites:
            print(i.getNom())
    def getHorizon(self):
        return self.Horizon
    def getDictParams(self):
        return self.DictParams
    def getListeActivitesSecondaires(self):
        return self.ListeActivitesSecondaires
    
class TableauSaisie:
    def __init__(self,Titre="",IntituleLigne="",IntituleColonne="",Taille = [1,1]):
        self.Titre = Titre
        self.IntituleLigne = IntituleLigne
        self.IntituleColonne = IntituleColonne
        self.Taille = Taille # [Indice Ligne , Indice Colonne]
        self.TableauAffichage = pd.DataFrame(np.zeros((Taille[0],Taille[1])))
        self.TableauDonnees = np.zeros((Taille[0],Taille[1]))
        if self.Taille[1] == 1:
            self.TableauAffichage.columns = [self.IntituleColonne]
        else:
            self.TableauAffichage.columns = [self.IntituleColonne+' '+str(i+1) for i in range(Taille[1])]
        if self.Taille[0] == 1:
            self.TableauAffichage.index = [self.IntituleLigne]
        else:
            self.TableauAffichage.index = [self.IntituleLigne+' '+str(i+1) for i in range(Taille[0])]
    
    #Setters
    def setTitre(self,Titre):
        self.Titre = Titre
    def setIntituleLignes(self,IntituleLigne):
        self.IntituleLigne = IntituleLigne
    def setIntituleColonne(self,IntituleColonne):
        self.IntituleColonne = IntituleColonne
    def setTaille(self,Taille):
        self.Taille = Taille
    #def setTableauAffichage(self,TableauAffichage):
        #self.TableauAffichage = TableauAffichage
    def setTableauDonnees(self,TableauDonnees):
        self.TableauDonnees = TableauDonnees
        self.TableauAffichage = pd.DataFrame(self.TableauDonnees)
        # self.TableauAffichage.columns = [self.IntituleColonne+' '+str(i+1) for i in range(self.Taille[1])]
        # self.TableauAffichage.index = [self.IntituleLigne+' '+str(i+1) for i in range(self.Taille[0])]
        if self.Taille[1] == 1:
            self.TableauAffichage.columns = [self.IntituleColonne]
        else:
            self.TableauAffichage.columns = [self.IntituleColonne+' '+str(i+1) for i in range(self.Taille[1])]
        if self.Taille[0] == 1:
            self.TableauAffichage.index = [self.IntituleLigne]
        else:
            self.TableauAffichage.index = [self.IntituleLigne+' '+str(i+1) for i in range(self.Taille[0])]
    
    #Getters
    def getTitre(self):
        return self.Titre    
    def getIntituleLigne(self):
        return self.IntituleLigne  
    def getIntituleColonne(self):
        return self.IntituleColonne  
    def getTaille(self):
        return self.Taille  
    def getTableauAffichage(self):
        return self.TableauAffichage  
    def getTableauDonnees(self):
        return self.TableauDonnees  
    
    
class Activite:
    def __init__(self,nom = "",listRev=[],listCout=[]):
        self.nom      = nom
        self.listRev  = listRev
        self.listCout = listCout
    
    #Setters
    def setNom(self,a) :
        self.nom = a
    def setlistRev(self,b) :
        self.listRev = b
    def setlistCout(self,c) :
        self.listCout = c
     
    #Getters
    def getNom(self) :
        return self.nom
    def getlistRev(self) :
        return self.listRev
    def getlistCout(self) :
        return self.listCout
    def showlistCout(self):
        for i in self.listCout:
            print(i.getNom())
    
    def IdentifierCouts(self):
        self.listCout = []
        print("Identification des Couts")
        wb = load_workbook('References.xlsx')
        ws = wb['Ref couts']
        for col in ws.iter_cols(min_row=2,min_col=4, max_col=ws.max_column, max_row=2):
            for cell in col:
                if (self.nom == cell.value):
                    coltosearch = cell.col_idx
        for row in ws.iter_rows(min_row=3,min_col=coltosearch, max_col=coltosearch, max_row=ws.max_row):
            for cell in row:
                if (cell.value == "x"):
                    C = Cout(ws['C'+str(cell.row)].value)
                    self.listCout.append(C)
                    
    
    def IdentifierRevenus(self):
        self.listRev = []
        print("Identification des Revenus")
        wb = load_workbook('References.xlsx')
        ws = wb['Ref revenus']
        for col in ws.iter_cols(min_row=2,min_col=3, max_col=ws.max_column, max_row=2):
            for cell in col:
                if (self.nom == cell.value):
                    coltosearch = cell.col_idx
        for row in ws.iter_rows(min_row=3,min_col=coltosearch, max_col=coltosearch, max_row=ws.max_row):
            for cell in row:
                if (cell.value == "x"):
                    R = Revenu(ws['A'+str(cell.row)].value)
                    self.listRev.append(R)
        
class Cpc:
    
    def __init__(self,ListeActivites = [],Dotation = [],ServiceDette = []) :
        self.ListeActivites    = ListeActivites
        self.Dotation          = Dotation
        self.ServiceDette      = ServiceDette
        
        
    #Setters
    def setListeActivites(self,a) :
        self.ListeActivites = a
    def setDotation(self,b) :
        self.Dotation = b
    def setServiceDette(self,c) :
        self.ServiceDette = c
     
    #Getters
    def getListeActivites(self) :
        return self.ListeActivites
    def getDotation(self) :
        return self.Dotation
    def getServiceDette(self) :
        return self.ServiceDette
    
#Definition des classes couts et activtés
class Cout:
    def __init__(self,Nom = "", Horizon=0,SaisieStartCol = 1,CoutCPC = []):
        self.Nom     = Nom
        self.Horizon = Horizon
        self.CoutCPC = CoutCPC
        self.listeTableaux = []
        self.listeTableauxMarche = []
        self.SaisieStartCol = SaisieStartCol
        self.DictTableaux = {}
        self.resultat = []
        self.isOnce = False
        self.step = ''
        self.VEchantillon = []
        self.DicoFormes = {}
        self.DicoFormesMarche = {}
        
    def resizeTableauxMarche(self):
        for titre in self.DicoFormesMarche.keys():
            self.listeTableauxMarche.append(TableauSaisie(titre,self.DicoFormesMarche[titre]['IntituleLigne'],self.DicoFormesMarche[titre]['IntituleColonne'],[self.DicoFormesMarche[titre]['Taille0'],self.DicoFormesMarche[titre]['Taille1']]))
    def resizeTableaux(self):
        for titre in self.DicoFormes.keys():
            self.listeTableaux.append(TableauSaisie(titre,self.DicoFormes[titre]['IntituleLigne'],self.DicoFormes[titre]['IntituleColonne'],[self.DicoFormes[titre]['Taille0'],self.DicoFormes[titre]['Taille1']]))
        
    #Setters
    def setNom(self, Nom):
        self.Nom = Nom
    def setHorizon(self,Horizon):
        self.Horizon = Horizon
    def setCoutCPC(self,CoutCPC):
        self.CoutCPC = CoutCPC
    def setSaisieStartCol(self,SaisieStartCol):
        self.SaisieStartCol = SaisieStartCol
    def setListTableauxMarche(self,listeTableauxMarche):
        self.listeTableauxMarche = listeTableauxMarche
    def setResultat(self,resultat):
        self.resultat = resultat
    
    #Getters
    def getNom(self):
        return self.Nom
    def getHorizon(self):
        return self.Horizon
    def getCoutCPC(self):
        return self.CoutCPC
    def getListTableaux(self):
        return self.listeTableaux
    def getSaisieStartCol(self):
        return self.SaisieStartCol
    def getListTableauxMarche(self):
        return self.listeTableauxMarche
    
    def SaisieIntrinseque(self):
        #Au préalable necessite le nom du cout, l'horizon, SaisieStartCol
        wb = load_workbook("References.xlsx")
        ws=wb['Cout x Tableau']
        for col in ws.iter_cols(min_row=2,min_col=3, max_col=ws.max_column, max_row=2):
            for cell in col:
                if (cell.value == self.Nom):
                    coltosearch = cell.col_idx
        #Liste temporaire occupant les titres des Tableaux de saisie
        Titres = []
        for row in ws.iter_rows(min_row=2,min_col=coltosearch, max_col=coltosearch, max_row=ws.max_row):
            for cell in row:
                if cell.value == 'x':
                    if ws['B'+str(cell.row)].value == 'int':
                        Titres.append(ws['A'+str(cell.row)].value)
        wsTableau = wb['Tableau x Features']
        for j in Titres:
            for col in wsTableau.iter_cols(min_row=2,min_col=2, max_col=ws.max_column, max_row=2):
                for cell in col:
                    if (cell.value == j):
                        Taille = [0,0]
                        IntituleLigne = wsTableau[cell.column+'3'].value
                        TypeCol = wsTableau[cell.column+'4'].value
                        IntituleColonne = wsTableau[cell.column+'5'].value
                        #Un Test sur le type d'indices ligne pour le tableau
                        if IntituleLigne == 'Horizon':
                            Taille[0] = self.Horizon
                        else:
                            Taille[0] = int(input("Veuillez saisir votre "+IntituleLigne))
                        if TypeCol == 1:
                            Taille[1] = 1
                        else:
                            Taille[1] = int(input("Veuillez saisir votre "+str(TypeCol)))
                        self.listeTableaux.append(TableauSaisie(j,IntituleLigne,IntituleColonne,Taille))
                        
    def SaisieIntrinseque_1(self):
        #Au préalable necessite le nom du cout, l'horizon, SaisieStartCol
        wb = load_workbook("References.xlsx")
        ws=wb['Cout x Tableau']
        for col in ws.iter_cols(min_row=2,min_col=3, max_col=ws.max_column, max_row=2):
            for cell in col:
                if (cell.value == self.Nom):
                    coltosearch = cell.col_idx
        #Liste temporaire occupant les titres des Tableaux de saisie
        Titres = []
        for row in ws.iter_rows(min_row=2,min_col=coltosearch, max_col=coltosearch, max_row=ws.max_row):
            for cell in row:
                if cell.value == "x" :
                    if ws['B'+str(cell.row)].value == 'int':
                        Titres.append(ws['A'+str(cell.row)].value)
                        print(Titres)
        wsTableau = wb['Tableau x Features'] 
        Formulaire = []
        for j in Titres:
            self.DicoFormes[j] ={}
            for col in wsTableau.iter_cols(min_row=2,min_col=2, max_col=ws.max_column, max_row=2):
                for cell in col:
                    if (cell.value == j):
                        Questions = [j]
                        Taille = [0,0]
                        self.DicoFormes[j]['Taille0'] = Taille[0]
                        self.DicoFormes[j]['Taille1'] = Taille[1]
                        IntituleLigne = wsTableau[cell.column+'3'].value
                        self.DicoFormes[j]['IntituleLigne'] = str(IntituleLigne)
                        TypeCol = wsTableau[cell.column+'4'].value
                        self.DicoFormes[j]['TypeCol'] = str(TypeCol)
                        IntituleColonne = wsTableau[cell.column+'5'].value
                        self.DicoFormes[j]['IntituleColonne'] = str(IntituleColonne)
                        #Un Test sur le type d'indices ligne pour le tableau
                        if IntituleLigne == 'Horizon':
                            Taille[0] = self.Horizon
                            self.DicoFormes[j]['IntituleLigne'] = self.pasVisualisation
                            self.DicoFormes[j]['Taille0'] = self.Horizon
                        elif type(IntituleLigne) == int:
                            Taille[0] = IntituleLigne
                            self.DicoFormes[j]['Taille0'] = IntituleLigne
                        else:
                            Questions.append(["Taille0","Veuillez saisir votre "+IntituleLigne])
                        if type(TypeCol) == int:
                            Taille[1] = TypeCol
                            self.DicoFormes[j]['Taille1'] = TypeCol
                        else:
                            Questions.append(["Taille1","Veuillez saisir votre "+str(TypeCol)])
                        Formulaire.append(Questions)
        return Formulaire
                        #self.listeTableaux.append(TableauSaisie(j,IntituleLigne,IntituleColonne,Taille))
        

    def SaisieMarche(self):
        wb = load_workbook("References.xlsx")
        ws=wb['Cout x Tableau']
        for col in ws.iter_cols(min_row=2,min_col=3, max_col=ws.max_column, max_row=2):
            for cell in col:
                if (cell.value == self.Nom):
                    coltosearch = cell.col_idx
        #Liste temporaire occupant les titres des Tableaux de saisie
        Titres = []
        for row in ws.iter_rows(min_row=2,min_col=coltosearch, max_col=coltosearch, max_row=ws.max_row):
            for cell in row:
                if cell.value == 'x':
                    if ws['B'+str(cell.row)].value == 'ext':
                        Titres.append(ws['A'+str(cell.row)].value)
        wsTableau = wb['Tableau x Features']
        for j in Titres:
            for col in wsTableau.iter_cols(min_row=2,min_col=2, max_col=ws.max_column, max_row=2):
                for cell in col:
                    if (cell.value == j):
                        Taille = [0,0]
                        IntituleLigne = wsTableau[cell.column+'3'].value
                        TypeCol = wsTableau[cell.column+'4'].value
                        IntituleColonne = wsTableau[cell.column+'5'].value
                        #Un Test sur le type d'indices ligne pour le tableau
                        if IntituleLigne == 'Horizon':
                            Taille[0] = self.Horizon
                        else:
                            Taille[0] = int(input("Veuillez saisir votre "+IntituleLigne))
                        if TypeCol == 1:
                            Taille[1] = 1
                        else:
                            Taille[1] = int(input("Veuillez saisir votre "+str(TypeCol)))
                        self.listeTableauxMarche.append(TableauSaisie(j,IntituleLigne,IntituleColonne,Taille))
                        
    def SaisieMarche_1(self):
        wb = load_workbook("References.xlsx")
        ws=wb['Cout x Tableau']
        for col in ws.iter_cols(min_row=2,min_col=3, max_col=ws.max_column, max_row=2):
            for cell in col:
                if (cell.value == self.Nom):
                    coltosearch = cell.col_idx
        #Liste temporaire occupant les titres des Tableaux de saisie
        Titres = []
        print(Titres)
        Formulaire = []
        for row in ws.iter_rows(min_row=2,min_col=coltosearch, max_col=coltosearch, max_row=ws.max_row):
            for cell in row:
                if cell.value == 'x':
                    if ws['B'+str(cell.row)].value == 'ext':
                        Titres.append(ws['A'+str(cell.row)].value)
                        print('ok')
        wsTableau = wb['Tableau x Features']
        for j in Titres:
            self.DicoFormesMarche[j] ={}
            for col in wsTableau.iter_cols(min_row=2,min_col=2, max_col=ws.max_column, max_row=2):
                for cell in col:
                    if (cell.value == j):
                        Questions = [j]
                        Taille = [0,0]
                        self.DicoFormesMarche[j]['Taille0'] = Taille[0]
                        self.DicoFormesMarche[j]['Taille1'] = Taille[1]
                        IntituleLigne = wsTableau[cell.column+'3'].value
                        self.DicoFormesMarche[j]['IntituleLigne'] = str(IntituleLigne)
                        TypeCol = wsTableau[cell.column+'4'].value
                        self.DicoFormesMarche[j]['TypeCol'] = str(TypeCol)
                        IntituleColonne = wsTableau[cell.column+'5'].value
                        self.DicoFormesMarche[j]['IntituleColonne'] = str(IntituleColonne)
                        #Un Test sur le type d'indices ligne pour le tableau
                        if IntituleLigne == 'Horizon':
                            Taille[0] = self.Horizon
                            self.DicoFormesMarche[j]['IntituleLigne'] = self.pasVisualisation
                            self.DicoFormesMarche[j]['Taille0'] = self.Horizon
                        elif type(IntituleLigne) == int:
                            Taille[0] = IntituleLigne
                            self.DicoFormesMarche[j]['Taille0'] = IntituleLigne
                        else:
                            Questions.append(["Taille0","Veuillez saisir votre "+IntituleLigne])
                        if type(TypeCol) == int:
                            Taille[1] = TypeCol
                            self.DicoFormesMarche[j]['Taille1'] = TypeCol
                        else:
                            Questions.append(["Taille1","Veuillez saisir votre "+str(TypeCol)])
                        Formulaire.append(Questions)
        return Formulaire
                        
    def getDicoTableaux(self):
        for i in self.listeTableaux:
            self.DictTableaux[i.getTitre()] = i.getTableauDonnees()
        for i in self.listeTableauxMarche:
            self.DictTableaux[i.getTitre()] = i.getTableauDonnees()
        return self.DictTableaux
        
    def CalculCout(self):
        CalculerCout(self)

class Revenu:
    def __init__(self,Nom = "", Horizon=0,SaisieStartCol = 1,CoutCPC = []):
        self.Nom     = Nom
        self.Horizon = Horizon
        self.RevenuCPC = CoutCPC
        self.listeTableaux = []
        self.listeTableauxMarche = []
        self.SaisieStartCol = SaisieStartCol
        self.DictTableaux = {}
        self.resultat = []
        self.Produits = {}
        self.isOnce = False
        self.step = ''
        self.VEchantillon = []
        self.DicoFormes = {}
        self.DicoFormesMarche = {}
        
    def resizeTableauxMarche(self):
        for titre in self.DicoFormesMarche.keys():
            self.listeTableauxMarche.append(TableauSaisie(titre,self.DicoFormesMarche[titre]['IntituleLigne'],self.DicoFormesMarche[titre]['IntituleColonne'],[self.DicoFormesMarche[titre]['Taille0'],self.DicoFormesMarche[titre]['Taille1']]))
    def resizeTableaux(self):
        for titre in self.DicoFormes.keys():
            self.listeTableaux.append(TableauSaisie(titre,self.DicoFormes[titre]['IntituleLigne'],self.DicoFormes[titre]['IntituleColonne'],[self.DicoFormes[titre]['Taille0'],self.DicoFormes[titre]['Taille1']]))
    #Setters
    def setNom(self, Nom):
        self.Nom = Nom
    def setHorizon(self,Horizon):
        self.Horizon = Horizon
    def setRevenuCPC(self,CoutCPC):
        self.RevenuCPC = CoutCPC
    def setSaisieStartCol(self,SaisieStartCol):
        self.SaisieStartCol = SaisieStartCol
    def setListTableauxMarche(self,listeTableauxMarche):
        self.listeTableauxMarche = listeTableauxMarche
    def setResultat(self,resultat):
        self.resultat = resultat
    def DicoTableaux(Self) :
        self.DicoTableaux = {}
    #Getters
    def getNom(self):
        return self.Nom
    def getHorizon(self):
        return self.Horizon
    def getRevenuCPC(self):
        return self.RevenuCPC
    def getListTableaux(self):
        return self.listeTableaux
    def getSaisieStartCol(self):
        return self.SaisieStartCol
    def getListTableauxMarche(self):
        return self.listeTableauxMarche
    
    def SaisieIntrinseque(self):
        #Au préalable necessite le nom du cout, l'horizon, SaisieStartCol
        wb = load_workbook("References.xlsx")
        ws=wb['Rev x Tableau']
        for col in ws.iter_cols(min_row=2,min_col=3, max_col=12, max_row=2):
            for cell in col:
                if (cell.value == self.Nom):
                    coltosearch = cell.col_idx
        #Liste temporaire occupant les titres des Tableaux de saisie
        Titres = []
        for row in ws.iter_rows(min_row=2,min_col=coltosearch, max_col=coltosearch, max_row=24):
            for cell in row:
                if cell.value == "x" :
                    if ws['B'+str(cell.row)].value == 'int':
                        Titres.append(ws['A'+str(cell.row)].value)
                        print(Titres)
        wsTableau = wb['Tableau x Features rev'] #@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
        for j in Titres:
            for col in wsTableau.iter_cols(min_row=2,min_col=2, max_col=23, max_row=2):
                for cell in col:
                    if (cell.value == j):
                        Taille = [0,0]
                        IntituleLigne = wsTableau[cell.column+'3'].value
                        TypeCol = wsTableau[cell.column+'4'].value
                        IntituleColonne = wsTableau[cell.column+'5'].value
                        #Un Test sur le type d'indices ligne pour le tableau
                        if IntituleLigne == 'Horizon':
                            Taille[0] = self.Horizon
                        else:
                            Taille[0] = int(input("Veuillez saisir votre "+IntituleLigne))
                        if TypeCol == 1:
                            Taille[1] = 1
                        else:
                            Taille[1] = int(input("Veuillez saisir votre "+str(TypeCol)))
                        self.listeTableaux.append(TableauSaisie(j,IntituleLigne,IntituleColonne,Taille))

    def SaisieIntrinseque_1(self):
        #Au préalable necessite le nom du cout, l'horizon, SaisieStartCol
        wb = load_workbook("References.xlsx")
        ws=wb['Rev x Tableau']
        for col in ws.iter_cols(min_row=2,min_col=3, max_col=ws.max_column, max_row=2):
            for cell in col:
                if (cell.value == self.Nom):
                    coltosearch = cell.col_idx
        #Liste temporaire occupant les titres des Tableaux de saisie
        Titres = []
        for row in ws.iter_rows(min_row=2,min_col=coltosearch, max_col=coltosearch, max_row=ws.max_row):
            for cell in row:
                if cell.value == "x" :
                    if ws['B'+str(cell.row)].value == 'int':
                        Titres.append(ws['A'+str(cell.row)].value)
                        print(Titres)
        wsTableau = wb['Tableau x Features rev'] #@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
        Formulaire = []
        for j in Titres:
            self.DicoFormes[j] ={}
            for col in wsTableau.iter_cols(min_row=2,min_col=2, max_col=ws.max_column, max_row=2):
                for cell in col:
                    if (cell.value == j):
                        Questions = [j]
                        Taille = [0,0]
                        self.DicoFormes[j]['Taille0'] = Taille[0]
                        self.DicoFormes[j]['Taille1'] = Taille[1]
                        IntituleLigne = wsTableau[cell.column+'3'].value
                        self.DicoFormes[j]['IntituleLigne'] = str(IntituleLigne)
                        TypeCol = wsTableau[cell.column+'4'].value
                        self.DicoFormes[j]['TypeCol'] = str(TypeCol)
                        IntituleColonne = wsTableau[cell.column+'5'].value
                        self.DicoFormes[j]['IntituleColonne'] = str(IntituleColonne)
                        #Un Test sur le type d'indices ligne pour le tableau
                        if IntituleLigne == 'Horizon':
                            Taille[0] = self.Horizon
                            self.DicoFormes[j]['IntituleLigne'] = self.pasVisualisation
                            self.DicoFormes[j]['Taille0'] = self.Horizon
                        elif IntituleLigne == 'Affluence':
                            Taille[0] = convertir_affluence(self.Horizon,self.pasVisualisation)
                            self.DicoFormes[j]['IntituleLigne'] = 'A'
                            self.DicoFormes[j]['Taille0'] = Taille[0]
                        elif type(IntituleLigne) == int:
                            Taille[0] = IntituleLigne
                            self.DicoFormes[j]['Taille0'] = IntituleLigne
                        else:
                            Questions.append(["Taille0","Veuillez saisir votre "+IntituleLigne])
                        if type(TypeCol) == int:
                            Taille[1] = TypeCol
                            self.DicoFormes[j]['Taille1'] = TypeCol
                        else:
                            Questions.append(["Taille1","Veuillez saisir votre "+str(TypeCol)])
                        Formulaire.append(Questions)
        return Formulaire
                        #self.listeTableaux.append(TableauSaisie(j,IntituleLigne,IntituleColonne,Taille))

    def SaisieMarche(self):
        wb = load_workbook("References.xlsx")
        ws=wb['Rev x Tableau']
        for col in ws.iter_cols(min_row=2,min_col=3, max_col=ws.max_column, max_row=2):
            for cell in col:
                if (cell.value == self.Nom):
                    coltosearch = cell.col_idx
        #Liste temporaire occupant les titres des Tableaux de saisie
        Titres = []
        print(Titres)
        for row in ws.iter_rows(min_row=2,min_col=coltosearch, max_col=coltosearch, max_row=ws.max_row):
            for cell in row:
                if cell.value == 'x':
                    if ws['B'+str(cell.row)].value == 'ext':
                        Titres.append(ws['A'+str(cell.row)].value)
                        print('ok')
        wsTableau = wb['Tableau x Features rev']
        for j in Titres:
            for col in wsTableau.iter_cols(min_row=2,min_col=2, max_col=23, max_row=2):
                for cell in col:
                    if (cell.value == j):
                        Taille = [0,0]
                        IntituleLigne = wsTableau[cell.column+'3'].value
                        TypeCol = wsTableau[cell.column+'4'].value
                        IntituleColonne = wsTableau[cell.column+'5'].value
                        #Un Test sur le type d'indices ligne pour le tableau
                        if IntituleLigne == 'Horizon':
                            Taille[0] = self.Horizon
                        else:
                            Taille[0] = int(input("Veuillez saisir votre "+IntituleLigne))
                        if TypeCol == 1 :
                            Taille[1] = 1
                        else:
                            Taille[1] = int(input("Veuillez saisir votre "+str(TypeCol)))
                        self.listeTableauxMarche.append(TableauSaisie(j,IntituleLigne,IntituleColonne,Taille))
    def SaisieMarche_1(self):
        wb = load_workbook("References.xlsx")
        ws=wb['Rev x Tableau']
        for col in ws.iter_cols(min_row=2,min_col=3, max_col=ws.max_column, max_row=2):
            for cell in col:
                if (cell.value == self.Nom):
                    coltosearch = cell.col_idx
        #Liste temporaire occupant les titres des Tableaux de saisie
        Titres = []
        print(Titres)
        Formulaire = []
        for row in ws.iter_rows(min_row=2,min_col=coltosearch, max_col=coltosearch, max_row=ws.max_row):
            for cell in row:
                if cell.value == 'x':
                    if ws['B'+str(cell.row)].value == 'ext':
                        Titres.append(ws['A'+str(cell.row)].value)
                        print('ok')
        wsTableau = wb['Tableau x Features rev']
        for j in Titres:
            self.DicoFormesMarche[j] ={}
            for col in wsTableau.iter_cols(min_row=2,min_col=2, max_col=ws.max_column, max_row=2):
                for cell in col:
                    if (cell.value == j):
                        Questions = [j]
                        Taille = [0,0]
                        self.DicoFormesMarche[j]['Taille0'] = Taille[0]
                        self.DicoFormesMarche[j]['Taille1'] = Taille[1]
                        IntituleLigne = wsTableau[cell.column+'3'].value
                        self.DicoFormesMarche[j]['IntituleLigne'] = str(IntituleLigne)
                        TypeCol = wsTableau[cell.column+'4'].value
                        self.DicoFormesMarche[j]['TypeCol'] = str(TypeCol)
                        IntituleColonne = wsTableau[cell.column+'5'].value
                        self.DicoFormesMarche[j]['IntituleColonne'] = str(IntituleColonne)
                        #Un Test sur le type d'indices ligne pour le tableau
                        if IntituleLigne == 'Horizon':
                            Taille[0] = self.Horizon
                            self.DicoFormesMarche[j]['IntituleLigne'] = self.pasVisualisation
                            self.DicoFormesMarche[j]['Taille0'] = self.Horizon
                        elif type(IntituleLigne) == int:
                            Taille[0] = IntituleLigne
                            self.DicoFormesMarche[j]['Taille0'] = IntituleLigne
                        else:
                            Questions.append(["Taille0","Veuillez saisir votre "+IntituleLigne])
                        if type(TypeCol) == int:
                            Taille[1] = TypeCol
                            self.DicoFormesMarche[j]['Taille1'] = TypeCol
                        else:
                            Questions.append(["Taille1","Veuillez saisir votre "+str(TypeCol)])
                        Formulaire.append(Questions)
        return Formulaire
        
    def getDicoTableaux(self):
        for i in self.listeTableaux:
            self.DictTableaux[i.getTitre()] = i.getTableauDonnees()
        for i in self.listeTableauxMarche:
            self.DictTableaux[i.getTitre()] = i.getTableauDonnees()
        return self.DictTableaux
        
    def CalculRev(self):
        CalculerRevenu(self)
        pass

class Dette:    #Nouvelle classe #### Standby
    def __init__(self, bailleur="", total=0, taux=0):
        self.periodicite = ' '   #LIER avec inputs intérface
        self.horizon = 0       #LIER avec inputs intérface
        self.bailleur = bailleur      #LIER avec inputs intérface
        self.total = total        #LIER avec inputs intérface
        self.taux = taux           #LIER avec inputs intérface
        self.date_debut_remboursement = 0 #LIER avec inputs interface
        self.apport_dette = [0]*self.horizon #taille horizon, déblocages pour chaque période
        self.amortissement_capital = [0]*self.horizon #taille horizon, capital à rembourser pour chaque période 



    def interets(self): #forme simple : penser à interfaçage avec moteur de calcul de Younes
        t = 0
        if self.periodicite == 'A'      : t = 1   #renvoie nbr de cases à regrouper     
        if self.periodicite == 'S'  : t = 2        
        if self.periodicite == 'T' : t = 3
        if self.periodicite == 'M'     : t = 12
        l = [0]*int((self.horizon/t))
        for i in range(int(self.date_debut_remboursement/t)+1,len(l)) :
            l[i] = self.taux*sum(self.amortissement_capital[i:])
        s = [0]*self.horizon
        c = 0
        for i in range(0,self.horizon,t):
            s[i] = l[c]
            c += 1
            
        return l    ### d = Dette(
    
def collect_financial_data(p): #executer par défaut 
    
    dictio = {}
    
    l = p.getListeActivites() + p.getListeActivitesSecondaires()
    for i in l :
        print(i.getNom())
        
    for i in l :
        dictio[i.getNom()] = { 'Couts': {} , 'Revenus': {} }
        for j in i.getlistCout() :
            dictio[i.getNom()]['Couts'][j.getNom()] = j.getresultat()
        for j in i.getlistRev() :
            dictio[i.getNom()]['Revenus'][j.getNom()] = j.getresultat()
    
    return dictio

    
def remplir_pour_test(p,h):    #rempli les vecteurs resultat pour tester
    for i in p.getListeActivites()+p.getListeActivitesSecondaires():
        for j in i.getlistRev():
            j.setResultat([0]*h)
            
        for j in i.getlistCout():
            j.setResultat([0]*h)   
    return 0        

def f(x):
    if       0 < x <= 300000  : return 0.1
    if  300000 < x <= 1000000 : return 0.2
    if 1000000 < x <= 5000000 : return 0.3
    if 5000000 < x            : return 0.31
    else                      : return 0
    
    
    
## RESULTAT D'EXPLOITATION
    


def tab_revenus(p): #renvoie tous les revenus /activite dans une liste de vecteurs : [Activité, p1,p2,...]
    l = []
    d = collect_financial_data(p)
    for i in d :
        for j in d[i]['Revenus']:
            l.append([j]+d[i]['Revenus'][j])
    ttc = l
    tva = l
    for i in ttc :
        for j in ttc[i]:
            if isinstance(ttc[i][j],float) == True :
                ttc[i][j] = (1 + TVA(p))*l[i][j]
    for i in tva :
        for j in tva[i]:
            if isinstance(tva[i][j],float) == True :
                tva[i][j] = TVA(p)*l[i][j]
    return (ttc,tva)
                    
def total_revenus(p): #renvoie un vecteur de taille horizon qui contient la somme des revenus
    d = collect_financial_data(p);
    h = p.getHorizon()
    l = [0]*(h+1)
    l[0] = "Total Revenus"
    c = 1
    while c < h+1 :
        for i in d.keys() :
            for j in d[i]['Revenus'].keys():
                l[c] += d[i]['Revenus'][j][c-1]
        c += 1  
    for i in range(1,h+1):
        l[i] = (1 + TVA(p))*l[i]
    return l[1:]
    
def tab_couts(p):  #renvoie tous les couts par activite dans une liste de vecteurs : [Activité, p1,p2,...]
    l = []
    d = collect_financial_data(p)
    for i in d :
        for j in d[i]['Couts']:
            l.append([j]+d[i]['Couts'][j])
    for i in l :
        for j in l[i]:
            if isinstance(l[i][j],float) == True :
                l[i][j] = (1 + taux(p))*l[i][j]
    return l
    
def total_couts(p): #renvoie un vecteur de taille horizon qui contient la somme des couts
    d = collect_financial_data(p);
    h = p.getHorizon()
    l = [0]*(h+1)
    l[0] = "Total Couts"
    c = 1
    while c < h+1 :
        for i in d.keys() :
            for j in d[i]['Couts'].keys():
                l[c] += d[i]['Couts'][j][c-1]    
        c += 1    
    return l
          
def resultat_d_exploitation(p):
    l = (p.getHorizon()+1)*[0]
    l[0] = "RESULTAT D'EXPLOITATION"
    r = total_revenus(p)
    c = total_couts(p)
    for i in range(1,p.getHorizon()+1):
        l[i] = r[i]+c[i]
    return l[1:]
    
## RESULTAT FINANCIER

def produit_financiers(p):
    return p.getproduits_financiers()

def charges_financieres(p):
    return interets(p)
    #return p.getcharges_financieres()
    
def amortissement_dettes(p): #les amortissements emprunteur /emprunteur + total
    l = []
    for i in p.getdette():
        inf = i.getdate_debut_remboursement()
        sup = i.getdate_maturite()
        t = (p.getHorizon()+1)*[0]
        t[0] = i.getbailleur()
        t[inf+1:sup+2]=i.getamortissement()
        l.append(t)
    s = [0]*(p.getHorizon()+1)
    s[0] = 'Total des amortissements'
    for j in range(1,p.getHorizon()+1):
        for i in l :
            s[j] += i[j]
    l.append(s[1:])
    return l        
        
        
        
def interets_dettes(p):  #les interets emprunteur /emprunteur + total
    l = []
    for  i in p.getdette():
        inf = i.getdate_debut_deblocage()
        t = (p.getHorizon()+1)*[0]
        t[inf:inf+i.getmaturite] = i.interets()
        t[0] = i.getbailleur()
        l.append(t)
    s = [0]*(p.getHorizon()+1)
    s[0] = 'Total des interets'
    for j in range(1,p.getHorizon()+1):
        for i in l :
            s[j] += i[j]
    l.append(s[1:])
    return l
    
def resultat_financier(p):
    l = ['RESULTAT FINANCIER'] + [0]*p.getHorizon()
    for i in range(1,p.getHorizon()+1):
        l[i] = p.getproduits_financiers()[i] - p.getcharges_financieres()[i]
    return l[1:]
    
    
## RESULTAT COURANT

def resultat_courant(p):
    l = ['RESULTAT COURANT'] + [0]*p.getHorizon()
    exp = resultat_d_exploitation(p)
    fin = resultat_financier(p)
    for i in range(1,p.getHorizon()+1):
        l[i] = exp[i] + fin[i] 
    return l[1:]

## RESULTAT NON COURANT    

def produits_non_courants(p):
    p.getproduits_non_courants()
    
def charges_non_courantes(p):
    p.getcharges_non_courantes()
 
def resultat_non_courant(p):
    l = []
    p = produits_non_courants(p)
    c = charges_non_courantes(p)
    for i in range(1,p.getHorizon()+1) :
        l[i] = p[i] - c[i] 
    return (['RESULTAT NON COURANT'] + l)[1:]
  
## RESULTAT AVANT IMPOTS

def resultat_avant_impot(p):
    l=['RESULTAT AVANT IMPOT'] + [0]*p.getHorizon() 
    courant  = resultat_courant(p)
    ncourant = resultat_non_courant(p)
    for i in range(1,p.getHorizon()):
        l[i] = resultat_courant(p)[i] + resultat_non_courant(p)[i]
    return l[1:]
  
## FISCALITE
#### Début modif 1 ####          
def IS(p):
    r = p.RAI
    per = p.pasVisualisation
    
    t = 0
    if per == 'A'      : t = 1   #renvoie nbr de cases à regrouper     
    if per == 'S'  : t = 2       
    if per == 'T' : t = 4
    if per == 'M'     : t = 12
    
    l = [0]*int(roundN(((len(r))/t)))
    for i in range(0,len(l)) :
        for j in range(i*t,min((i+1)*t,len(r))):
            l[i] += r[j]
        # for j in range(0,t) :
            # l[i] += r[i*t+j]
    
    s = l.copy()
    for i in range(0,len(l)):
        if l[i] < 0  : 
            s[i] = l[i]*f(l[i])
        else :
            c = l[i]
            n=0
            if i >=4:
                n = 4
            else:
                n=i
            for j in range(i-n,i):
                print(l[j])
                if l[j] < 0 :
                    if c > - l[j]  :
                        c = c + l[j]
                        l[j] = float('inf')
                    else :
                        tmp = c
                        c = 0
                        l[j] = l[j] + tmp
                
                    
            s[i] = c*f(c)
                    
    impots = [0]*len(p.RAI)
    for i in range(0,len(s)):
        impots[min(((i+1)*t)-1,len(impots)-1)] = s[i]
    return(impots)

    
##### Fin Modif 1 #######
def IS_ajustealaperiodicite(p,per):
    
    IS = IS(p)
    l = ['IS'] + [0]*p.getHorizon()
    
    t = 0
    if per == 'Annuelle'      : return IS       
    if per == 'Semestrielle'  : t = 2       #renvoie la période à laquelle il faut remettre l'IS
    if per == 'Trimestrielle' : t = 4
    if per == 'Mensuelle'     : t = 12
    
    c = 1
    for i in range(t,len(IS),t-1):
        l[i] = IS[c]
        c += 1
        
    return l[1:]
        
        
    
    
    
    
    
def TVA(p):
    s = p.getNom()
    taux = 0
    if s == "Ecole"                    : taux = 0.2
    if s == "Promotion immobiliere"    : taux = 0.2
    if s == "Restaurant"               : taux = 0.1
    if s == "Industrie manufacturiere" : taux = 0.2
    return taux
    
## RESULTAT NET

def resultat_NET(p):
    pass
    
def roundN(p):
    if p != int(p):
        return int(p) +1
    else:
        return p
        
def convertir_affluence(horizon,pas):
    t=0
    if pas == "M":  t=12
    if pas == "T":  t=4
    if pas == "S":  t=2
    if pas == "A":  t=1
    return roundN(horizon/t)
        